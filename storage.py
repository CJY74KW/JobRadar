"""
storage.py
분석 결과 저장 모듈
- Excel (시트별 분리): 전체 공고, 추천 공고, 마감 임박, 기술 통계, 지원 준비
- CSV: 전체 공고 백업
- SQLite: 누적 공고 DB 저장
"""

import os
import csv
import sqlite3
import json
from collections import Counter
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


JOB_COLS = ["source", "company", "title", "location", "deadline", "dday",
            "score", "match_detail", "tech_keywords", "link"]


def save_all(jobs: list[dict], recommended: list[dict],
             deadline_soon: list[dict], tech_stats: list[tuple],
             config: dict) -> None:
    output_dir = os.path.dirname(config["output"]["excel_path"])
    os.makedirs(output_dir, exist_ok=True)

    sorted_jobs = _sort_jobs_by_score(jobs)
    save_csv(sorted_jobs, config["output"]["csv_path"])
    save_all_jobs_excel(sorted_jobs, _get_all_excel_path(config), tech_stats)
    save_excel(sorted_jobs, recommended, deadline_soon, tech_stats, config)
    save_sqlite(sorted_jobs, config["output"]["db_path"])
    print(f"\n[저장 완료] 결과가 '{output_dir}' 폴더에 저장되었습니다.")


# ============================================================
# CSV 저장
# ============================================================
def save_csv(jobs: list[dict], path: str) -> None:
    if not jobs:
        return
    flat = [_flatten(j) for j in jobs]
    fieldnames = list(flat[0].keys())
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(flat)
    print(f"  CSV 저장: {path} ({len(jobs)}개)")


def save_all_jobs_excel(jobs: list[dict], path: str,
                        tech_stats: list[tuple] | None = None) -> None:
    """전체 채용공고를 적합도 점수 내림차순으로 저장한다."""
    output_dir = os.path.dirname(path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    sorted_jobs = _sort_jobs_by_score(jobs)
    df_all = _jobs_to_df(sorted_jobs, JOB_COLS, include_rank=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="전체공고", index=False)

    _append_all_jobs_summary(path, sorted_jobs, tech_stats)
    _apply_excel_style(path)
    print(f"  전체공고 Excel 저장: {path} ({len(jobs)}개, 점수순)")


# ============================================================
# Excel 저장
# ============================================================
def save_excel(jobs: list[dict], recommended: list[dict],
               deadline_soon: list[dict], tech_stats: list[tuple],
               config: dict) -> None:
    path = config["output"]["excel_path"]
    sorted_jobs = _sort_jobs_by_score(jobs)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # 시트1: 전체 공고
        df_all = _jobs_to_df(sorted_jobs, JOB_COLS, include_rank=True)
        df_all.to_excel(writer, sheet_name="전체공고", index=False)

        # 시트2: 추천 공고 (점수 상위)
        df_rec = _jobs_to_df(_sort_jobs_by_score(recommended), JOB_COLS, include_rank=True)
        df_rec.to_excel(writer, sheet_name="추천공고", index=False)

        # 시트3: 마감 임박 공고
        df_dl = _jobs_to_df(_sort_jobs_by_score(deadline_soon), JOB_COLS, include_rank=True)
        df_dl.to_excel(writer, sheet_name="마감임박", index=False)

        # 시트4: 기술 키워드 통계
        if tech_stats:
            df_tech = pd.DataFrame(tech_stats, columns=["기술키워드", "공고수"])
            df_tech.to_excel(writer, sheet_name="기술키워드통계", index=False)

        # 시트5: 지원 준비 (자소서 포인트 + 면접 질문)
        prep_rows = []
        for job in recommended[:10]:
            prep_rows.append({
                "회사명": job.get("company", ""),
                "공고명": job.get("title", ""),
                "마감일": job.get("deadline", ""),
                "자기소개서 작성 포인트": "\n".join(
                    f"• {p}" for p in job.get("cover_letter_points", [])
                ),
                "예상 면접 질문": "\n".join(
                    f"Q{i+1}. {q}" for i, q in enumerate(job.get("interview_questions", []))
                ),
            })
        if prep_rows:
            df_prep = pd.DataFrame(prep_rows)
            df_prep.to_excel(writer, sheet_name="지원준비", index=False)

    _apply_excel_style(path)
    print(f"  Excel 저장: {path}")


def _jobs_to_df(jobs: list[dict], cols: list[str],
                include_rank: bool = False) -> pd.DataFrame:
    rows = []
    for idx, job in enumerate(jobs, 1):
        row = {}
        if include_rank:
            row["rank"] = idx
        for col in cols:
            val = job.get(col, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            row[col] = val
        rows.append(row)
    col_rename = {
        "source": "출처", "company": "회사명", "title": "공고명",
        "location": "근무지", "deadline": "마감일", "dday": "D-Day",
        "score": "적합도점수", "match_detail": "매칭근거",
        "tech_keywords": "기술키워드", "link": "공고링크",
        "rank": "순위",
    }
    df_cols = (["rank"] if include_rank else []) + cols
    df = pd.DataFrame(rows, columns=df_cols)
    df.rename(columns=col_rename, inplace=True)
    return df


def _apply_excel_style(path: str) -> None:
    """헤더 강조 및 열 너비 자동 조정"""
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for ws in wb.worksheets:
        # 헤더 스타일
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border

        # 열 너비 자동 조정
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max(
                (len(str(c.value)) if c.value else 0 for c in col), default=0
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        # 지원준비 시트는 행 높이 크게
        if ws.title == "지원준비":
            for row in ws.iter_rows(min_row=2):
                ws.row_dimensions[row[0].row].height = 120
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)


def _append_all_jobs_summary(path: str, jobs: list[dict],
                             tech_stats: list[tuple] | None) -> None:
    """전체공고 시트 하단에 기술 키워드 현황과 대시보드용 차트를 추가한다."""
    wb = load_workbook(path)
    ws = wb["전체공고"]

    tech_rows = _normalise_stat_rows(tech_stats, limit=15)
    if not tech_rows:
        tech_rows = _get_tech_stats_from_jobs(jobs, limit=15)
    score_rows = _get_score_distribution(jobs)
    source_rows = _get_source_stats(jobs)

    start_row = ws.max_row + 3
    _write_section_title(ws, start_row, "기술 키워드 수요 현황")
    tech_header_row = start_row + 1
    tech_last_row = _write_table(ws, tech_header_row, ["기술키워드", "공고수"], tech_rows)
    if tech_rows:
        _add_bar_chart(
            ws,
            title="기술 키워드 TOP 15",
            header_row=tech_header_row,
            last_row=tech_last_row,
            anchor=f"D{start_row}",
            horizontal=True,
        )

    dashboard_row = max(tech_last_row + 3, start_row + 22)
    _write_section_title(ws, dashboard_row, "대시보드")

    score_header_row = dashboard_row + 1
    score_last_row = _write_table(ws, score_header_row, ["점수구간", "공고수"], score_rows)
    if score_rows:
        _add_bar_chart(
            ws,
            title="적합도 점수 분포",
            header_row=score_header_row,
            last_row=score_last_row,
            anchor=f"D{dashboard_row}",
        )

    source_header_row = max(score_last_row + 3, dashboard_row + 18)
    source_last_row = _write_table(ws, source_header_row, ["수집출처", "공고수"], source_rows)
    if source_rows:
        _add_pie_chart(
            ws,
            title="수집 출처별 공고 수",
            header_row=source_header_row,
            last_row=source_last_row,
            anchor=f"D{source_header_row - 1}",
        )

    wb.save(path)


def _write_section_title(ws, row: int, title: str) -> None:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=14, color="1F2937")
    cell.alignment = Alignment(horizontal="left")


def _write_table(ws, start_row: int, headers: list[str],
                 rows: list[tuple]) -> int:
    header_fill = PatternFill("solid", fgColor="4F6BED")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_idx, row_values in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")

    return start_row + len(rows)


def _add_bar_chart(ws, title: str, header_row: int, last_row: int,
                   anchor: str, horizontal: bool = False) -> None:
    if last_row <= header_row:
        return

    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    chart.style = 10
    chart.title = title
    chart.height = 8.5
    chart.width = 16
    chart.y_axis.majorGridlines = None if horizontal else chart.y_axis.majorGridlines

    data = Reference(ws, min_col=2, min_row=header_row, max_row=last_row)
    categories = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, anchor)


def _add_pie_chart(ws, title: str, header_row: int, last_row: int,
                   anchor: str) -> None:
    if last_row <= header_row:
        return

    chart = PieChart()
    chart.title = title
    chart.height = 8
    chart.width = 12

    data = Reference(ws, min_col=2, min_row=header_row + 1, max_row=last_row)
    categories = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data)
    chart.set_categories(categories)
    ws.add_chart(chart, anchor)


def _normalise_stat_rows(stats: list[tuple] | None,
                         limit: int | None = None) -> list[tuple]:
    rows = []
    for name, count in stats or []:
        if not name:
            continue
        try:
            count_value = int(count)
        except (TypeError, ValueError):
            count_value = 0
        rows.append((str(name), count_value))
    return rows[:limit] if limit else rows


def _get_tech_stats_from_jobs(jobs: list[dict], limit: int = 15) -> list[tuple]:
    counter = Counter()
    for job in jobs:
        keywords = job.get("tech_keywords", [])
        if isinstance(keywords, str):
            keywords = [kw.strip() for kw in keywords.split(",") if kw.strip()]
        for keyword in keywords:
            keyword = str(keyword).strip()
            if keyword:
                counter[keyword] += 1
    return counter.most_common(limit)


def _get_score_distribution(jobs: list[dict]) -> list[tuple]:
    bins = [("0-20", 0), ("21-40", 0), ("41-60", 0), ("61-80", 0), ("81-100", 0)]
    for job in jobs:
        score = _safe_score(job.get("score", 0))
        if score <= 20:
            idx = 0
        elif score <= 40:
            idx = 1
        elif score <= 60:
            idx = 2
        elif score <= 80:
            idx = 3
        else:
            idx = 4
        label, count = bins[idx]
        bins[idx] = (label, count + 1)
    return bins


def _get_source_stats(jobs: list[dict]) -> list[tuple]:
    counter = Counter(job.get("source", "기타") or "기타" for job in jobs)
    return counter.most_common()


def _get_all_excel_path(config: dict) -> str:
    configured = config["output"].get("all_excel_path", "")
    if configured:
        return configured
    csv_path = config["output"]["csv_path"]
    base, _ = os.path.splitext(csv_path)
    return f"{base}.xlsx"


def _sort_jobs_by_score(jobs: list[dict]) -> list[dict]:
    return sorted(jobs, key=lambda job: _safe_score(job.get("score", 0)), reverse=True)


def _safe_score(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


# ============================================================
# SQLite 저장
# ============================================================
def save_sqlite(jobs: list[dict], db_path: str) -> None:
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT,
            company TEXT,
            title TEXT,
            location TEXT,
            job_type TEXT,
            experience TEXT,
            deadline TEXT,
            dday TEXT,
            salary TEXT,
            score REAL,
            tech_keywords TEXT,
            link TEXT,
            collected_at TEXT,
            UNIQUE(company, title, deadline)
        )
    """)

    inserted = 0
    for job in jobs:
        try:
            cur.execute("""
                INSERT OR IGNORE INTO jobs
                (source, company, title, location, job_type, experience,
                 deadline, dday, salary, score, tech_keywords, link, collected_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                job.get("source", ""),
                job.get("company", ""),
                job.get("title", ""),
                job.get("location", ""),
                job.get("job_type", ""),
                job.get("experience", ""),
                job.get("deadline", ""),
                job.get("dday", ""),
                job.get("salary", ""),
                job.get("score", 0),
                ", ".join(job.get("tech_keywords", [])),
                job.get("link", ""),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ))
            if cur.rowcount:
                inserted += 1
        except sqlite3.Error:
            continue

    conn.commit()
    conn.close()
    print(f"  SQLite 저장: {db_path} (신규 {inserted}개)")


def _flatten(job: dict) -> dict:
    flat = {}
    for k, v in job.items():
        if isinstance(v, list):
            flat[k] = ", ".join(str(i) for i in v)
        elif isinstance(v, dict):
            flat[k] = json.dumps(v, ensure_ascii=False)
        else:
            flat[k] = v
    return flat
